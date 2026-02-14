"""
Import/Export Excel (XLSX) — массовая загрузка и выгрузка данных.
Поддерживает: компоненты, директивы, персонал ПЛГ, дефекты.
"""
import io
import logging
import uuid
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.api.deps import get_db, get_current_user
from app.api.helpers import audit

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/import-export", tags=["import-export"])


@router.get("/export/{entity_type}")
def export_xlsx(entity_type: str, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """
    Экспорт в XLSX. Типы: components, directives, bulletins, specialists, defects, work_orders.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    data_map = {
        "components": ("airworthiness_core", "_components", ["name", "part_number", "serial_number", "ata_chapter", "manufacturer", "condition", "current_hours", "current_cycles"]),
        "directives": ("airworthiness_core", "_directives", ["number", "title", "issuing_authority", "effective_date", "compliance_type", "status"]),
        "bulletins": ("airworthiness_core", "_bulletins", ["number", "title", "manufacturer", "category", "issued_date", "status"]),
        "specialists": ("personnel_plg", "_specialists", ["full_name", "personnel_number", "position", "category", "license_number", "status"]),
        "defects": ("defects", "_defects", ["aircraft_reg", "ata_chapter", "description", "severity", "discovered_during", "status"]),
        "work_orders": ("work_orders", "_work_orders", ["wo_number", "aircraft_reg", "wo_type", "title", "priority", "status", "estimated_manhours"]),
    }

    if entity_type not in data_map:
        raise HTTPException(400, f"Unknown entity: {entity_type}. Supported: {list(data_map.keys())}")

    module_name, store_name, columns = data_map[entity_type]
    mod = __import__(f"app.api.routes.{module_name}", fromlist=[store_name])
    items = list(getattr(mod, store_name).values())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = entity_type
    ws.append(columns)

    for item in items:
        row = [str(item.get(col, "")) for col in columns]
        ws.append(row)

    # Style header
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    audit(db, user, "export_xlsx", entity_type, description=f"XLSX export: {entity_type} ({len(items)} rows)")
    db.commit()

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={entity_type}_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


@router.post("/import/{entity_type}")
async def import_xlsx(
    entity_type: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Импорт из XLSX. Первая строка — заголовки."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Only .xlsx files accepted")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "File must have header + at least 1 data row")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    imported = 0
    errors = []

    for i, row in enumerate(rows[1:], start=2):
        try:
            item = {headers[j]: (str(v).strip() if v is not None else "") for j, v in enumerate(row) if j < len(headers)}
            item["id"] = str(uuid.uuid4())
            item["created_at"] = datetime.now(timezone.utc).isoformat()

            if entity_type == "components":
                from app.api.routes.airworthiness_core import _components
                if not item.get("name") or not item.get("part_number") or not item.get("serial_number"):
                    errors.append(f"Row {i}: missing required fields (name, part_number, serial_number)")
                    continue
                item.setdefault("condition", "serviceable")
                item["current_hours"] = float(item.get("current_hours", 0) or 0)
                item["current_cycles"] = int(item.get("current_cycles", 0) or 0)
                _components[item["id"]] = item
            elif entity_type == "specialists":
                from app.api.routes.personnel_plg import _specialists
                if not item.get("full_name") or not item.get("personnel_number"):
                    errors.append(f"Row {i}: missing full_name or personnel_number")
                    continue
                item.setdefault("status", "active")
                item.setdefault("specializations", [])
                _specialists[item["id"]] = item
            elif entity_type == "directives":
                from app.api.routes.airworthiness_core import _directives
                if not item.get("number") or not item.get("title"):
                    errors.append(f"Row {i}: missing number or title")
                    continue
                item.setdefault("status", "open")
                item.setdefault("aircraft_types", [])
                _directives[item["id"]] = item
            else:
                errors.append(f"Import not supported for: {entity_type}")
                break

            imported += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    audit(db, user, "import_xlsx", entity_type, description=f"XLSX import: {entity_type} ({imported} imported, {len(errors)} errors)")
    db.commit()

    return {"imported": imported, "errors": errors[:20], "total_rows": len(rows) - 1}
