import csv
import io
import zipfile
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, Depends, File, UploadFile, HTTPException
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.api.deps import get_current_user, require_roles
from app.db.session import get_db
from app.models import IngestJobLog, MaintenanceTask, DefectReport, LimitedLifeComponent, LandingGearComponent, ChecklistItem, ChecklistTemplate, Aircraft

router = APIRouter(tags=["ingest"])


class IngestLogCreate(BaseModel):
    source_system: str
    job_name: str
    status: str
    details: str | None = None


class ParseResultItem(BaseModel):
    name: str
    headers: list[str]
    rows: list[dict[str, Any]]
    row_count: int


class ParseArchiveResponse(BaseModel):
    items: list[ParseResultItem]
    errors: list[str] = []


class ImportTableRequest(BaseModel):
    target: str  # maintenance_tasks | defect_reports | limited_life_components | landing_gear_components | checklist_items
    aircraft_id: str | None = None
    template_id: str | None = None
    column_mapping: dict[str, str]  # {"field_name": "header_name"}
    rows: list[dict[str, Any]]


@router.post(
    "/ingest/logs",
    dependencies=[Depends(require_roles("admin", "authority_inspector"))],
)
def create_ingest_log(payload: IngestLogCreate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    log = IngestJobLog(
        source_system=payload.source_system,
        job_name=payload.job_name,
        status=payload.status,
        details=payload.details,
        started_at=datetime.now(timezone.utc),
        finished_at=datetime.now(timezone.utc),
    )
    db.add(log)
    db.commit()
    db.refresh(log)
    return log


@router.get(
    "/ingest/logs",
    dependencies=[Depends(require_roles("admin", "authority_inspector"))],
)
def list_ingest_logs(db: Session = Depends(get_db), user=Depends(get_current_user)):
    return db.query(IngestJobLog).order_by(IngestJobLog.created_at.desc()).limit(200).all()


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    """Парсит CSV файл, возвращает заголовки и строки."""
    try:
        text = content.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = content.decode('cp1251')
    
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows = list(reader)
    return headers, rows


def _parse_xlsx(content: bytes) -> list[tuple[str, list[str], list[dict[str, Any]]]]:
    """Парсит XLSX файл, возвращает список (sheet_name, headers, rows)."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    result = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row == 0:
            continue
        headers = [str(cell.value or '') for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append({headers[i]: str(val) if val is not None else '' for i, val in enumerate(row)})
        result.append((sheet_name, headers, rows))
    return result


@router.post(
    "/ingest/parse-archive",
    response_model=ParseArchiveResponse,
    dependencies=[Depends(require_roles("admin", "authority_inspector", "operator_manager"))],
)
async def parse_archive(
    file: UploadFile = File(...),
    user=Depends(get_current_user),
):
    """Парсит ZIP/CSV/XLSX архив и возвращает структурированные данные для табличного просмотра."""
    content = await file.read()
    items = []
    errors = []
    
    try:
        if file.filename.endswith('.zip'):
            with zipfile.ZipFile(io.BytesIO(content)) as z:
                for name in z.namelist():
                    if name.endswith('/'):
                        continue
                    try:
                        file_content = z.read(name)
                        if name.endswith('.csv'):
                            headers, rows = _parse_csv(file_content)
                            items.append(ParseResultItem(
                                name=name,
                                headers=headers,
                                rows=rows[:500],  # Ограничение для preview
                                row_count=len(rows)
                            ))
                        elif name.endswith(('.xlsx', '.xls')):
                            for sheet_name, headers, rows in _parse_xlsx(file_content):
                                items.append(ParseResultItem(
                                    name=f"{name}/{sheet_name}",
                                    headers=headers,
                                    rows=rows[:500],
                                    row_count=len(rows)
                                ))
                    except Exception as e:
                        errors.append(f"Ошибка обработки {name}: {str(e)}")
        elif file.filename.endswith('.csv'):
            headers, rows = _parse_csv(content)
            items.append(ParseResultItem(
                name=file.filename,
                headers=headers,
                rows=rows[:500],
                row_count=len(rows)
            ))
        elif file.filename.endswith(('.xlsx', '.xls')):
            for sheet_name, headers, rows in _parse_xlsx(content):
                items.append(ParseResultItem(
                    name=f"{file.filename}/{sheet_name}",
                    headers=headers,
                    rows=rows[:500],
                    row_count=len(rows)
                ))
        else:
            raise HTTPException(status_code=400, detail="Неподдерживаемый формат файла")
    except Exception as e:
        errors.append(f"Ошибка парсинга: {str(e)}")
    
    return ParseArchiveResponse(items=items, errors=errors)


@router.post(
    "/ingest/import-table",
    dependencies=[Depends(require_roles("admin", "authority_inspector", "operator_manager"))],
)
def import_table(
    payload: ImportTableRequest,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Импортирует данные из таблицы в указанную целевую таблицу."""
    imported = 0
    errors = []
    
    if payload.target == "maintenance_tasks":
        if not payload.aircraft_id:
            raise HTTPException(status_code=400, detail="aircraft_id обязателен для maintenance_tasks")
        aircraft = db.query(Aircraft).filter(Aircraft.id == payload.aircraft_id).first()
        if not aircraft:
            raise HTTPException(status_code=404, detail="ВС не найдено")
        
        for row in payload.rows:
            try:
                task = MaintenanceTask(aircraft_id=payload.aircraft_id)
                for field, header in payload.column_mapping.items():
                    if hasattr(task, field) and header in row:
                        setattr(task, field, row[header] or None)
                db.add(task)
                imported += 1
            except Exception as e:
                errors.append(f"Строка {imported + len(errors) + 1}: {str(e)}")
    
    elif payload.target == "defect_reports":
        if not payload.aircraft_id:
            raise HTTPException(status_code=400, detail="aircraft_id обязателен для defect_reports")
        for row in payload.rows:
            try:
                report = DefectReport(aircraft_id=payload.aircraft_id)
                for field, header in payload.column_mapping.items():
                    if hasattr(report, field) and header in row:
                        setattr(report, field, row[header] or None)
                db.add(report)
                imported += 1
            except Exception as e:
                errors.append(f"Строка {imported + len(errors) + 1}: {str(e)}")
    
    elif payload.target == "limited_life_components":
        if not payload.aircraft_id:
            raise HTTPException(status_code=400, detail="aircraft_id обязателен для limited_life_components")
        for row in payload.rows:
            try:
                comp = LimitedLifeComponent(aircraft_id=payload.aircraft_id)
                for field, header in payload.column_mapping.items():
                    if hasattr(comp, field) and header in row:
                        setattr(comp, field, row[header] or None)
                db.add(comp)
                imported += 1
            except Exception as e:
                errors.append(f"Строка {imported + len(errors) + 1}: {str(e)}")
    
    elif payload.target == "landing_gear_components":
        if not payload.aircraft_id:
            raise HTTPException(status_code=400, detail="aircraft_id обязателен для landing_gear_components")
        for row in payload.rows:
            try:
                comp = LandingGearComponent(aircraft_id=payload.aircraft_id)
                for field, header in payload.column_mapping.items():
                    if hasattr(comp, field) and header in row:
                        setattr(comp, field, row[header] or None)
                db.add(comp)
                imported += 1
            except Exception as e:
                errors.append(f"Строка {imported + len(errors) + 1}: {str(e)}")
    
    elif payload.target == "checklist_items":
        if not payload.template_id:
            raise HTTPException(status_code=400, detail="template_id обязателен для checklist_items")
        template = db.query(ChecklistTemplate).filter(ChecklistTemplate.id == payload.template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        
        for row in payload.rows:
            try:
                item = ChecklistItem(template_id=payload.template_id)
                for field, header in payload.column_mapping.items():
                    if hasattr(item, field) and header in row:
                        setattr(item, field, row[header] or None)
                db.add(item)
                imported += 1
            except Exception as e:
                errors.append(f"Строка {imported + len(errors) + 1}: {str(e)}")
    
    else:
        raise HTTPException(status_code=400, detail=f"Неподдерживаемый target: {payload.target}")
    
    db.commit()
    
    return {
        "imported": imported,
        "errors": errors,
        "status": "partial" if errors else "success"
    }
